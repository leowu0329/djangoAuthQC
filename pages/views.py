import openpyxl
from django.shortcuts import render, redirect, get_object_or_404
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from datetime import datetime,timedelta
from django.views.generic import TemplateView

# 匯入資料庫模型與表單
from .models import DefectList, OperatorList, OrderList, SpecList, IPQCList
from .forms import IPQCForm 

@login_required
def home_page(request):
    return render(request, 'home.html')

@login_required
def ipqc_list_view(request):
    # 1. 取得所有原始巡檢資料
    queryset = IPQCList.objects.all().order_by('-date', '-time')

    # 2. 取得前端傳過來的搜尋與日期篩選參數
    search_query = request.GET.get('search', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    # 3. 日期範圍篩選
    if start_date:
        queryset = queryset.filter(date__gte=start_date)
    if end_date:
        queryset = queryset.filter(date__lte=end_date)

    # 4. 關鍵字全欄位搜尋（支援空白隔開多條件查詢）
    if search_query:
        # 將輸入的字串依據空白（空格）切分成關鍵字列表，自動忽略重複與多餘的空格
        keywords = search_query.split()
        
        # 建立一個基礎的 Q 物件，用於串接多個關鍵字
        combined_q = Q()
        
        for word in keywords:
            # 針對「單一個關鍵字」，定義要在哪些欄位進行搜尋（欄位間是 OR | 關係）
            word_q = Q(order_number__icontains=word) | \
                     Q(operator__icontains=word) | \
                     Q(draw_ver__icontains=word) | \
                     Q(product_number__icontains=word) | \
                     Q(product_name__icontains=word) | \
                     Q(spec__icontains=word) | \
                     Q(inspector__icontains=word) | \
                     Q(defect_type__icontains=word) | \
                     Q(defect_status__icontains=word) | \
                     Q(handing_measure__icontains=word) | \
                     Q(mark__icontains=word)
            
            # 使用 AND (&=) 將每個單字的 Q 物件串接起來
            # 確保資料必須同時符合「關鍵字1」且符合「關鍵字2」...
            combined_q &= word_q
            
        # 執行多條件過濾
        queryset = queryset.filter(combined_q)

    # 5. 分頁處理（每頁固定顯示 10 筆記錄）
    paginator = Paginator(queryset, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # 6. 初始化一個空白的 Django 表單，供前端新增與編輯的 Modal 使用
    form = IPQCForm()

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'start_date': start_date,
        'end_date': end_date,
        'form': form,
    }
    return render(request, 'app/ipqc/ipqc_list.html', context)


# ==================== 處理新增紀錄 ====================
@login_required
def ipqc_create_view(request):
    if request.method == 'POST':
        form = IPQCForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '【成功】巡檢紀錄已成功新增！')
        else:
            error_msg = "【失敗】新增失敗，請檢查欄位格式："
            for field, errors in form.errors.items():
                error_msg += f" [{form.fields[field].label}: {', '.join(errors)}]"
            messages.error(request, error_msg)
            
    return redirect('ipqc_list')


# ==================== 修正焦點：精確定義 ipqc_edit_view 函數 ====================
@login_required
def ipqc_edit_view(request, pk):
    """
    編輯修改特定的 IPQC 巡檢紀錄
    """
    # 1. 修正 Model 大小寫：改為 IPQCList
    record = get_object_or_404(IPQCList, pk=pk)
    
    if request.method == 'POST':
        # 傳入 instance=record 執行 UPDATE
        form = IPQCForm(request.POST, request.FILES, instance=record)
        
        if form.is_valid():
            try:
                form.save()
                messages.success(request, f"巡檢紀錄修改成功！")
                # 2. 修正導向名稱，改回與常規列表一致的 'ipqc_list'
                return redirect('ipqc_list')
            except Exception as e:
                messages.error(request, f"儲存失敗，錯誤原因：{str(e)}")
        else:
            # 組合表單驗證失敗的詳細原因並透過 message 傳回
            error_msg = "【失敗】修改失敗，請檢查格式："
            for field, errors in form.errors.items():
                error_msg += f" [{form.fields[field].label}: {', '.join(errors)}]"
            messages.error(request, error_msg)
            return redirect('ipqc_list')
            
    return redirect('ipqc_list')


# ==================== 修正焦點：精確定義 ipqc_delete_view 函數 ====================
@login_required
def ipqc_delete_view(request, pk):
    if request.method == 'POST':
        instance = get_object_or_404(IPQCList, pk=pk)
        order_num = instance.order_number or "未填工單"
        instance.delete()
        messages.success(request, f'【成功】製令工單「{order_num}」的巡檢紀錄已成功刪除！')
    else:
        messages.error(request, '【失敗】無效的刪除請求方式。')
        
    return redirect('ipqc_list')


@login_required
def export_ipqc_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"IPQC_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IPQC 巡檢紀錄"
    
    # 完整 14 個欄位標頭
    headers = [
        "日期", "時間", "製令工單", "操作人員", "版本", 
        "品號", "品名", "規格", "數量", "檢驗人員", "判定",
        "不良類型", "不良狀態", "後續處置", "備註"
    ]
    ws.append(headers)
    
    # 依日期與時間排序匯出所有資料
    for item in IPQCList.objects.all().order_by('-date', '-time'):
        ws.append([
            item.date.strftime('%Y-%m-%d') if item.date else '',
            item.time.strftime('%H:%M:%S') if item.time else '',
            item.order_number or '',
            item.operator or '',
            item.draw_ver or '',
            item.product_number or '',
            item.product_name or '',
            item.spec or '',
            item.quantity if item.quantity is not None else 0,
            item.inspector or '',
            item.determination or '',
            item.defect_type or '',
            item.defect_status or '',
            item.handing_measure or '',
            item.mark or ''
        ])
        
    wb.save(response)
    return response

@login_required
def import_ipqc_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            
            success_count = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # 如果整列都是空的，或是關鍵的日期與工單皆為空，則跳過
                if not row[0] and not row[2]: 
                    continue
                
                # 日期型態安全轉換
                date_val = row[0]
                date_obj = None
                if date_val:
                    if isinstance(date_val, str):
                        date_obj = datetime.strptime(date_val.strip(), '%Y-%m-%d').date()
                    else:
                        date_obj = date_val

                # 時間型態安全轉換
                time_val = row[1]
                time_obj = None
                if time_val:
                    if isinstance(time_val, str):
                        time_obj = datetime.strptime(time_val.strip(), '%H:%M:%S').time()
                    else:
                        time_obj = time_val

                # 嚴格對齊 14 個欄位索引 (0 到 13)
                IPQCList.objects.create(
                    date=date_obj,
                    time=time_obj,
                    order_number=str(row[2]).strip() if row[2] is not None else row[2],
                    operator=str(row[3]).strip() if row[3] is not None else row[3],
                    draw_ver=str(row[4]).strip() if row[4] is not None else row[4],
                    product_number=str(row[5]).strip() if row[5] is not None else row[5],
                    product_name=str(row[6]).strip() if row[6] is not None else row[6],
                    spec=str(row[7]).strip() if row[7] is not None else row[7],
                    quantity=int(row[8]) if row[8] is not None else 0,
                    inspector=str(row[9]).strip() if row[9] is not None else row[9],
                    determination=str(row[10]).strip() if row[10] is not None else row[10],
                    defect_type=str(row[11]).strip() if row[11] is not None else row[11],
                    defect_status=str(row[12]).strip() if row[12] is not None else row[12],
                    handing_measure=str(row[13]).strip() if row[13] is not None else row[13],
                    mark=str(row[14]).strip() if row[14] is not None else row[14]
                )
                success_count += 1
                
            messages.success(request, f'Excel 資料匯入成功！共匯入 {success_count} 筆巡檢紀錄。')
        except Exception as e:
            messages.error(request, f'匯入失敗，錯誤原因: {str(e)}')
            
    return redirect('ipqc_list')

@login_required
def defect_list_view(request):
    return render(request, 'app/ipqc/defect_list.html', {'items': DefectList.objects.all()})

@login_required
def operator_list_view(request):
    return render(request, 'app/ipqc/operator_list.html', {'items': OperatorList.objects.all()})

@login_required
def order_list_view(request):
    return render(request, 'app/ipqc/order_list.html', {'items': OrderList.objects.all()})

@login_required
def spec_list_view(request):
    return render(request, 'app/ipqc/spec_list.html', {'items': SpecList.objects.all()})

@login_required
def ipqc_analyze_view(request):
    """
    品質圖表綜合分析視圖
    - Tab 1: 近三個月品質趨勢圖與細項統計表
    - Tab 2: 近三個月五大製程檢驗數/不良數/不良率表格 (修復)1
    - Tab 3: 2026年 1-9月 五大分類製程不良率推移圖與報表
    """
    # ==========================================
    # 【通用時間基準】設定為 2026 年 7 月
    # ==========================================
    current_date = datetime(2026, 7, 1) 
    
    # 五大製程基本配置
    process_configs = {
        'pump': {'name': '泵浦', 'prefix': '5341'},
        'machining': {'name': '機加', 'prefix': '5346'},
        'injection_add': {'name': '射加', 'prefix': '5348'},
        'injection': {'name': '射出', 'prefix': '5347'},
        'insert': {'name': '燒嵌', 'prefix': '5345'},
    }

    # ==========================================
    # 【Tab 1 & Tab 2: 近三個月數據處理】
    # ==========================================
    months_list_3m = []
    for i in range(2, -1, -1):
        check_date = current_date - timedelta(days=i*30)
        months_list_3m.append((check_date.year, check_date.month))
    
    labels_3m = []
    defect_batches_3m = []
    defect_rates_3m = []
    monthly_queries_3m = [] 

    for year, month in months_list_3m:
        labels_3m.append(f"{year}年{month}月")
        monthly_qs = IPQCList.objects.filter(date__year=year, date__month=month)
        monthly_queries_3m.append(monthly_qs)
        
        total_inspections = monthly_qs.count()
        bad_batches = monthly_qs.exclude(defect_type__isnull=True).exclude(defect_type="").exclude(defect_type="無").count()
        
        rate = 0.0
        if total_inspections > 0:
            rate = round((bad_batches / total_inspections) * 100, 2)
            
        defect_batches_3m.append(bad_batches)
        defect_rates_3m.append(rate)

    # Tab 1 下方表格：不良種類細項
    three_months_qs = IPQCList.objects.filter(
        Q(date__year=months_list_3m[0][0], date__month=months_list_3m[0][1]) |
        Q(date__year=months_list_3m[1][0], date__month=months_list_3m[1][1]) |
        Q(date__year=months_list_3m[2][0], date__month=months_list_3m[2][1])
    ).exclude(defect_type__isnull=True).exclude(defect_type="").exclude(defect_type="無")
    
    distinct_defects = list(three_months_qs.values_list('defect_type', flat=True).distinct())

    table_rows_3m = []
    for defect in distinct_defects:
        counts_by_month = []
        for monthly_qs in monthly_queries_3m:
            count = monthly_qs.filter(defect_type=defect).count()
            counts_by_month.append(count)
        table_rows_3m.append({
            'defect_type': defect,
            'counts': counts_by_month
        })

    # Tab 2 表格：修復製程不良率統計表 (橫軸為月份，縱軸為製程)
    process_table_data = []
    for key, config in process_configs.items():
        row_data = {'name': config['name'], 'prefix': config['prefix'], 'months': []}
        for monthly_qs in monthly_queries_3m:
            proc_qs = monthly_qs.filter(order_number__startswith=config['prefix'])
            inspections = proc_qs.count()
            defects = proc_qs.exclude(defect_type__isnull=True).exclude(defect_type="").exclude(defect_type="無").count()
            rate = round((defects / inspections) * 100, 2) if inspections > 0 else 0.0
            row_data['months'].append({
                'inspections': inspections,
                'defects': defects,
                'rate': rate
            })
        process_table_data.append(row_data)

    # ==========================================
    # 【Tab 3: 不良分類推移圖】數據處理 (1月 - 9月)
    # ==========================================
    labels_9m = [f"{m}月" for m in range(1, 10)]
    trend_matrix_9m = []

    for key, config in process_configs.items():
        rates_by_month = []
        for month in range(1, 10):
            monthly_process_qs = IPQCList.objects.filter(
                order_number__startswith=config['prefix'],
                date__year=2026,
                date__month=month
            )
            inspections_cnt = monthly_process_qs.count()
            defects_cnt = monthly_process_qs.exclude(defect_type__isnull=True).exclude(defect_type="").exclude(defect_type="無").count()
            fail_rate = round((defects_cnt / inspections_cnt) * 100, 2) if inspections_cnt > 0 else 0.0
            rates_by_month.append(fail_rate)
            
        trend_matrix_9m.append({
            'name': config['name'],
            'prefix': config['prefix'],
            'rates': rates_by_month
        })

    context = {
        # Tab 1 數據
        'labels': labels_3m,
        'defect_batches': defect_batches_3m,
        'defect_rates': defect_rates_3m,
        'table_rows': table_rows_3m,
        
        # Tab 2 數據 (修復成功)
        'process_table_data': process_table_data,
        
        # Tab 3 數據
        'labels_9m': labels_9m,
        'trend_matrix_9m': trend_matrix_9m,
    }
    return render(request, 'app/ipqc/ipqc_analyze.html', context)